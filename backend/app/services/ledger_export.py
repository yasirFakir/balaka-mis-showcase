from datetime import datetime
from typing import List, Optional, Dict, Any
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.models.transaction import Transaction
from app.models.service_request import ServiceRequest
from app.models.vendor import VendorTransaction, Vendor
from app.models.service import ServiceDefinition
from app.models.system import SystemSetting
from io import BytesIO
import csv

# For Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

# For PDF
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
except ImportError:
    SimpleDocTemplate = None

class LedgerExportService:
    def __init__(self, db: Session):
        self.db = db

    def generate_export(
        self,
        start_date: datetime,
        end_date: datetime,
        format_type: str,
        components: List[str],
        currency: str = "SAR",
        scope: str = "both",
        service_ids: Optional[List[int]] = None,
        sort_order: str = "desc"
    ) -> Any:
        # 1. Fetch Data
        data = self._fetch_data(start_date, end_date, components, scope, currency, service_ids, sort_order)
        
        # 2. Generate File
        if format_type.lower() == "excel":
            return self._generate_excel(data, start_date, end_date, currency)
        elif format_type.lower() == "pdf":
            return self._generate_pdf(data, start_date, end_date, currency)
        else:
            raise ValueError("Unsupported format")

    def _get_rate(self, target_currency: str) -> float:
        if target_currency == "SAR":
            return 1.0
            
        # 1. Check for manual override
        manual_enabled = self.db.query(SystemSetting).filter(SystemSetting.key == "currency_manual_enabled").first()
        if manual_enabled and manual_enabled.value_bool:
             manual_rate = self.db.query(SystemSetting).filter(SystemSetting.key == "currency_manual_rate").first()
             return manual_rate.value_float if manual_rate else 32.0
             
        # 2. Try Fallback / External
        try:
            import requests
            resp = requests.get("https://open.er-api.com/v6/latest/SAR", timeout=3.0)
            if resp.status_code == 200:
                api_data = resp.json()
                return api_data.get("rates", {}).get(target_currency, 32.0)
        except Exception as e:
            print(f"Export Rate Fetch Error: {e}")
            pass
            
        return 32.0 # Hard Fallback

    def _fetch_data(self, start_date: datetime, end_date: datetime, components: List[str], scope: str, currency: str, service_ids: Optional[List[int]] = None, sort_order: str = "desc") -> Dict[str, Any]:
        result = {}
        rate = self._get_rate(currency)
        result["meta"] = {"rate": rate}
        
        from sqlalchemy import desc, asc
        from sqlalchemy.orm import joinedload
        sort_func = desc if sort_order == "desc" else asc

        # Base filters
        req_filters = [
            ServiceRequest.created_at >= start_date,
            ServiceRequest.created_at <= end_date,
            ServiceRequest.status != "Cancelled"
        ]
        
        # Scope Filter: Internal vs Public
        if scope == "internal":
            req_filters.append(ServiceRequest.service_definition.has(ServiceDefinition.is_public == False))
        elif scope == "public":
            req_filters.append(ServiceRequest.service_definition.has(ServiceDefinition.is_public == True))
            
        # Service ID Filter
        if service_ids and len(service_ids) > 0:
            req_filters.append(ServiceRequest.service_def_id.in_(service_ids))
        
        # Transactions also need filtering by scope via ServiceRequest join
        txn_filters = [
            Transaction.created_at >= start_date,
            Transaction.created_at <= end_date,
            Transaction.status == "Verified"
        ]
        if scope != "both":
            txn_filters.append(Transaction.service_request.has(ServiceRequest.service_definition.has(ServiceDefinition.is_public == (scope == "public"))))
            
        if service_ids and len(service_ids) > 0:
            txn_filters.append(Transaction.service_request.has(ServiceRequest.service_def_id.in_(service_ids)))

        
        if "transaction_log" in components or "all" in components:
            txns = self.db.query(Transaction).options(
                joinedload(Transaction.user),
                joinedload(Transaction.service_request).joinedload(ServiceRequest.service_definition)
            ).filter(*txn_filters).order_by(sort_func(Transaction.created_at)).all()
            result["transactions"] = txns

        if "revenue_stream" in components or "all" in components:
            revenue = self.db.query(
                ServiceRequest.service_def_id,
                func.count(ServiceRequest.id).label("count"),
                func.sum(ServiceRequest.selling_price).label("total_revenue")
            ).join(ServiceDefinition).filter(*req_filters).group_by(ServiceRequest.service_def_id).all()
            
            enriched_revenue = []
            for r in revenue:
                svc = self.db.query(ServiceDefinition).get(r.service_def_id)
                total_rev = (r.total_revenue or 0.0) * rate
                enriched_revenue.append({
                    "service": svc.name if svc else "Unknown",
                    "count": r.count,
                    "total": total_rev
                })
            result["revenue"] = enriched_revenue

        if "vendor_payment" in components or "all" in components:
            vendor_txns = self.db.query(VendorTransaction).join(Vendor).filter(
                VendorTransaction.created_at >= start_date,
                VendorTransaction.created_at <= end_date,
                VendorTransaction.transaction_type == "PAYMENT"
            ).order_by(sort_func(VendorTransaction.created_at)).all()
            result["vendor_payments"] = vendor_txns

        if "profitability" in components or "all" in components:
            profit_stats = self.db.query(
                func.sum(ServiceRequest.profit).label("total_profit"),
                func.sum(ServiceRequest.selling_price).label("total_sales"),
                func.sum(ServiceRequest.cost_price).label("total_cost")
            ).join(ServiceDefinition).filter(*req_filters).first()
            
            total_sales = (profit_stats.total_sales or 0.0) * rate
            total_cost = (profit_stats.total_cost or 0.0) * rate
            total_profit = (profit_stats.total_profit or 0.0) * rate
            
            result["profitability"] = {
                "total_profit": total_profit,
                "total_sales": total_sales,
                "total_cost": total_cost,
                "margin": (total_profit / total_sales) * 100 if total_sales else 0
            }
            
            # Fetch detailed Sales Log for "detailed info" request
            sales_log = self.db.query(ServiceRequest).join(ServiceDefinition).filter(*req_filters).order_by(sort_func(ServiceRequest.created_at)).all()
            result["sales_log"] = sales_log

        return result

    def _generate_excel(self, data: Dict, start_date, end_date, currency):
        if not Workbook:
            raise ImportError("openpyxl is not installed")
        
        try:
            wb = Workbook()
            rate = data["meta"]["rate"]
            
            # Gonia Theme Colors (Invoice Style)
            ABYSS_BLUE = "065084"
            MIDNIGHT_VIOLET = "1E0741"
            HORIZON_TEAL = "78B9B5"
            WHITE = "FFFFFF"
            
            # Styling Constants
            header_font = Font(bold=True, color=WHITE)
            header_fill = PatternFill(start_color=MIDNIGHT_VIOLET, end_color=MIDNIGHT_VIOLET, fill_type="solid")
            
            total_font = Font(bold=True, color=WHITE)
            total_fill = PatternFill(start_color=HORIZON_TEAL, end_color=HORIZON_TEAL, fill_type="solid")
            
            section_title_font = Font(size=12, bold=True, color=ABYSS_BLUE)
            
            # Helper to style a row as a header
            def apply_header_style(ws, row_idx):
                for cell in ws[row_idx]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

            # Helper to style a row as a total
            def apply_total_style(ws, row_idx):
                for cell in ws[row_idx]:
                    cell.font = total_font
                    cell.fill = total_fill
                    cell.alignment = Alignment(horizontal="right")

            # Helper to adjust column widths
            def adjust_column_widths(ws):
                for column in ws.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    if adjusted_width > 50: adjusted_width = 50
                    ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

            # --- SHEET 1: SUMMARY ---
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            ws_summary.append(["Balaka MIS - Financial Ledger Summary"])
            ws_summary['A1'].font = Font(size=16, bold=True, color=ABYSS_BLUE)
            ws_summary.append([f"Period: {start_date.date()} to {end_date.date()}"])
            ws_summary.append([f"Reporting Currency: {currency} (Rate: {rate:.4f})"])
            ws_summary.append([f"Generated On: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
            ws_summary.append([])

            if "profitability" in data:
                p = data["profitability"]
                ws_summary.append(["PROFITABILITY OVERVIEW"])
                ws_summary[f'A{ws_summary.max_row}'].font = section_title_font
                ws_summary.append(["Metric", f"Value ({currency})"])
                apply_header_style(ws_summary, ws_summary.max_row)
                ws_summary.append(["Total Sales", round(p["total_sales"], 2)])
                ws_summary.append(["Total Cost", round(p["total_cost"], 2)])
                ws_summary.append(["Net Profit", round(p["total_profit"], 2)])
                ws_summary.append(["Margin %", f"{p['margin']:.2f}%"])
                ws_summary.append([])

            if "revenue" in data:
                ws_summary.append(["REVENUE STREAM BY SERVICE"])
                ws_summary[f'A{ws_summary.max_row}'].font = section_title_font
                ws_summary.append(["Service Name", "Request Count", f"Total Revenue ({currency})"])
                apply_header_style(ws_summary, ws_summary.max_row)
                total_rev_sum = 0
                for r in data["revenue"]:
                    ws_summary.append([r["service"], r["count"], round(r["total"], 2)])
                    total_rev_sum += r["total"]
                
                # Total Row
                ws_summary.append(["TOTAL REVENUE", "", round(total_rev_sum, 2)])
                apply_total_style(ws_summary, ws_summary.max_row)
            
            adjust_column_widths(ws_summary)

            # --- SHEET 2: TRANSACTION LOG ---
            if "transactions" in data:
                ws_txns = wb.create_sheet("Transaction Log")
                ws_txns.append(["Date", "User ID", "User Name", "Service", "TXN ID", "Ex. Rate", f"Amount ({currency})", "Status", "Method", "Notes"])
                apply_header_style(ws_txns, 1)
                
                total_txn_amount = 0
                for t in data["transactions"]:
                    amount = t.amount * rate
                    user_id = f"#{t.user_id}" if t.user_id else "-"
                    user_name = t.user.full_name if t.user else "-"
                    service_name = t.service_request.service_definition.name if t.service_request and t.service_request.service_definition else "-"
                    rate_label = f"{t.exchange_rate:.2f} ({t.claimed_currency or 'SAR'}/SAR)"
                    
                    ws_txns.append([
                        t.created_at.strftime("%Y-%m-%d"),
                        user_id,
                        user_name,
                        service_name,
                        t.transaction_id,
                        rate_label,
                        round(amount, 2),
                        t.status,
                        t.payment_method,
                        t.notes
                    ])
                    total_txn_amount += amount
                
                # Total Row
                ws_txns.append(["TOTAL LOGGED", "", "", "", "", "", round(total_txn_amount, 2), "", "", ""])
                apply_total_style(ws_txns, ws_txns.max_row)
                adjust_column_widths(ws_txns)

            # --- SHEET 3: VENDOR PAYMENTS ---
            if "vendor_payments" in data:
                ws_vendors = wb.create_sheet("Vendor Payments")
                ws_vendors.append(["Date", "Vendor ID", "Vendor Name", "TXN ID", f"Amount ({currency})", "Notes"])
                apply_header_style(ws_vendors, 1)
                
                total_vendor_amount = 0
                for v in data["vendor_payments"]:
                    amount = v.amount * rate
                    vendor_id = f"#{v.vendor_id}" if v.vendor_id else "-"
                    vendor_name = v.vendor.name if v.vendor else "Unknown"
                    
                    ws_vendors.append([
                        v.created_at.strftime("%Y-%m-%d"),
                        vendor_id,
                        vendor_name,
                        v.transaction_id,
                        round(amount, 2),
                        v.notes
                    ])
                    total_vendor_amount += amount
                
                # Total Row
                ws_vendors.append(["TOTAL PAYMENTS", "", "", "", round(total_vendor_amount, 2), ""])
                apply_total_style(ws_vendors, ws_vendors.max_row)
                adjust_column_widths(ws_vendors)

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output
        except Exception as e:
            error_out = BytesIO()
            error_out.write(f"Excel Generation Failed: {str(e)}".encode('utf-8'))
            error_out.seek(0)
            return error_out

    def _generate_pdf(self, data: Dict, start_date, end_date, currency):
        if not SimpleDocTemplate:
            raise ImportError("reportlab is not installed")

        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        rate = data["meta"]["rate"]
        
        # Invoice Theme Colors
        ABYSS_BLUE = colors.HexColor("#065084")
        MIDNIGHT_VIOLET = colors.HexColor("#1E0741")
        HORIZON_TEAL = colors.HexColor("#78B9B5")
        
        # Custom Title Style
        title_style = styles['Title']
        title_style.textColor = ABYSS_BLUE
        title_style.alignment = 0 # Left
        
        # Header
        elements.append(Paragraph("Air Balaka International // Financial Registry", title_style))
        elements.append(Paragraph(f"Period: {start_date.date()} to {end_date.date()}", styles['Normal']))
        elements.append(Paragraph(f"Reporting Currency: {currency} (Exchange Rate: {rate:.2f})", styles['Normal']))
        elements.append(Spacer(1, 24))
        
        # Profitability
        if "profitability" in data:
            p = data["profitability"]
            elements.append(Paragraph("I. Profitability Overview", styles['Heading2']))
            prof_data = [
                ["Metric", f"Value ({currency})"],
                ["Total Sales", f"{p['total_sales']:.2f}"],
                ["Total Cost", f"{p['total_cost']:.2f}"],
                ["Net Profit", f"{p['total_profit']:.2f}"],
                ["Margin %", f"{p['margin']:.2f}%"]
            ]
            t = Table(prof_data, colWidths=[200, 200])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), MIDNIGHT_VIOLET),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                # Style last row as total if it were a sum, but here it's just margin
                ('BACKGROUND', (0, -1), (-1, -1), HORIZON_TEAL),
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 12))

        # Revenue
        if "revenue" in data:
            elements.append(Paragraph("II. Revenue Stream by Service", styles['Heading2']))
            rev_data = [["Service", "Request Count", f"Total ({currency})"]]
            total_rev_sum = 0
            for r in data["revenue"]:
                rev_data.append([r["service"], str(r["count"]), f"{r['total']:.2f}"])
                total_rev_sum += r["total"]
            
            rev_data.append(["TOTAL REVENUE", "", f"{total_rev_sum:.2f}"])
            
            t = Table(rev_data, colWidths=[250, 100, 150])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), MIDNIGHT_VIOLET),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                # Total Row
                ('BACKGROUND', (0, -1), (-1, -1), HORIZON_TEAL),
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('ALIGN', (0, -1), (-1, -1), 'RIGHT'),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 12))

        # Transactions
        if "transactions" in data:
            elements.append(Paragraph("III. Transaction Log", styles['Heading2']))
            txn_data = [["Date", "User ID", "Name", "TXN ID", "Rate", f"Amount ({currency})", "Status"]]
            total_txn_amount = 0
            for tx in data["transactions"][:200]: 
                amount = tx.amount * rate
                user_id = f"#{tx.user_id}" if tx.user_id else "-"
                user_name = tx.user.full_name if tx.user else "-"
                rate_label = f"{tx.exchange_rate:.2f} ({tx.claimed_currency or 'SAR'}/SAR)"
                
                txn_data.append([
                    tx.created_at.strftime("%Y-%m-%d"),
                    user_id,
                    user_name,
                    tx.transaction_id,
                    rate_label,
                    f"{amount:.2f}",
                    tx.status
                ])
                total_txn_amount += amount
            
            if len(data["transactions"]) > 200:
                txn_data.append(["...", "...", "...", "...", "...", "...", "..."])
            
            txn_data.append(["TOTAL LOGGED", "", "", "", "", f"{total_txn_amount:.2f}", ""])
                
            t = Table(txn_data, colWidths=[70, 60, 120, 100, 80, 80, 80])
            t.setStyle(TableStyle([
                 ('BACKGROUND', (0, 0), (-1, 0), MIDNIGHT_VIOLET),
                 ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                 ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                 ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                 ('FONTSIZE', (0, 0), (-1, -1), 8),
                 # Total Row
                 ('BACKGROUND', (0, -1), (-1, -1), HORIZON_TEAL),
                 ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
                 ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                 ('ALIGN', (0, -1), (-1, -1), 'RIGHT'),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 12))

        # Vendor Payments
        if "vendor_payments" in data:
            elements.append(Paragraph("IV. Vendor Payments", styles['Heading2']))
            vnd_data = [["Date", "Vendor ID", "Vendor Name", "TXN ID", f"Amount ({currency})", "Notes"]]
            total_vendor_amount = 0
            for v in data["vendor_payments"][:200]:
                amount = v.amount * rate
                vendor_id = f"#{v.vendor_id}" if v.vendor_id else "-"
                vendor_name = v.vendor.name if v.vendor else "Unknown"
                vnd_data.append([
                    v.created_at.strftime("%Y-%m-%d"),
                    vendor_id,
                    vendor_name,
                    v.transaction_id,
                    f"{amount:.2f}",
                    v.notes[:50] + "..." if v.notes and len(v.notes) > 50 else (v.notes or "-")
                ])
                total_vendor_amount += amount
            
            if len(data["vendor_payments"]) > 200:
                vnd_data.append(["...", "...", "...", "...", "...", "..."])
            
            vnd_data.append(["TOTAL PAYMENTS", "", "", "", f"{total_vendor_amount:.2f}", ""])

            t = Table(vnd_data, colWidths=[70, 60, 150, 100, 80, 200])
            t.setStyle(TableStyle([
                 ('BACKGROUND', (0, 0), (-1, 0), MIDNIGHT_VIOLET),
                 ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                 ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                 ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                 ('FONTSIZE', (0, 0), (-1, -1), 8),
                 # Total Row
                 ('BACKGROUND', (0, -1), (-1, -1), HORIZON_TEAL),
                 ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
                 ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                 ('ALIGN', (0, -1), (-1, -1), 'RIGHT'),
            ]))
            elements.append(t)

        doc.build(elements)
        output.seek(0)
        return output